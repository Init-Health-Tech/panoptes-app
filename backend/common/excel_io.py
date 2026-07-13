"""Excel template + bulk import helpers (openpyxl)."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="1B4F72")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXAMPLE_FILL = PatternFill("solid", fgColor="EAF2F8")


def workbook_response(workbook: Workbook, filename: str) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_template(
    *,
    sheet_title: str,
    headers: list[str],
    example_row: list[Any] | None = None,
    dropdowns: dict[str, list[str]] | None = None,
    notes: list[str] | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(header) + 4)

    if example_row:
        for col, value in enumerate(example_row, start=1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.fill = EXAMPLE_FILL

    if dropdowns:
        for header, choices in dropdowns.items():
            if header not in headers:
                continue
            col = headers.index(header) + 1
            letter = get_column_letter(col)
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(choices) + '"',
                allow_blank=True,
            )
            dv.error = "Valor no permitido"
            dv.errorTitle = "Validación"
            ws.add_data_validation(dv)
            start_row = 2 if example_row else 2
            dv.add(f"{letter}{start_row}:{letter}2000")

    if notes:
        notes_ws = wb.create_sheet("Instrucciones")
        notes_ws["A1"] = "Instrucciones de carga masiva"
        notes_ws["A1"].font = Font(bold=True, size=14)
        for idx, note in enumerate(notes, start=3):
            notes_ws.cell(row=idx, column=1, value=note)
        notes_ws.column_dimensions["A"].width = 100

    return wb


def read_sheet_rows(uploaded_file, *, expected_headers: list[str]) -> tuple[list[str], list[dict[str, Any]], list[dict]]:
    """
    Returns (headers, rows_as_dicts, parse_errors).
    Row dicts include `_row` (1-based Excel row number).
    """
    errors: list[dict] = []
    try:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [], [{"row": 0, "field": "file", "message": f"No se pudo leer el Excel: {exc}"}]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [], [{"row": 1, "field": "headers", "message": "El archivo está vacío."}]

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    missing = [h for h in expected_headers if h not in headers]
    if missing:
        errors.append(
            {
                "row": 1,
                "field": "headers",
                "message": f"Faltan columnas obligatorias: {', '.join(missing)}. "
                f"Descarga la plantilla oficial.",
            }
        )
        return headers, [], errors

    parsed: list[dict[str, Any]] = []
    for excel_row, values in enumerate(rows_iter, start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        row: dict[str, Any] = {"_row": excel_row}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = values[idx] if idx < len(values) else None
            if isinstance(value, str):
                value = value.strip()
            row[header] = value
        parsed.append(row)

    return headers, parsed, errors


def cell_str(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def cell_bool(row: dict, key: str, default: bool = False) -> bool:
    value = cell_str(row, key).lower()
    if not value:
        return default
    return value in {"1", "true", "si", "sí", "yes", "y", "x"}
